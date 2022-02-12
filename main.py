import smtplib
import imaplib
import speech_recognition as sr
import pyttsx3
import email
from email.message import EmailMessage
from email.header import decode_header
import openpyxl as xl
import joblib
from kivymd.app import MDApp
from kivy.core.window import Window
from kivy.lang.builder import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.uix.list import TwoLineAvatarListItem, IconLeftWidget, OneLineAvatarIconListItem
import copy

# for listener to work install pyaudio package

email_receivers = []
addresses = []
sub = []
body = []
new_contacts = []
spam_senders = []
contact_icons = {'a': 'A.png', 'b': 'B.png', 'c': 'C.png', 'd': 'D.png', 'e': 'E.png', 'f': 'F.png',
                 'g': 'G.png', 'h': 'H.png', 'i': 'I.png', 'j': 'J.png', 'k': 'K.png', 'l': 'L.png',
                 'm': 'M.png', 'n': 'N.png', 'o': 'O.png', 'p': 'P.png', 'q': 'Q.png', 'r': 'R.png',
                 's': 'S.png', 't': 'T.png', 'u': 'U.png', 'v': 'V.png', 'w': 'W.png', 'x': 'X.png',
                 'y': 'Y.png', 'z': 'Z.png'}

wb = xl.load_workbook('contacts.xlsx')
sheet = wb['Sheet1']
contact_list = {}
x = 2

for x in range(2, sheet.max_row + 1):
    cell1 = sheet.cell(x, 2)
    cell2 = sheet.cell(x, 3)
    contact_list[cell1.value] = cell2.value

cv = joblib.load('vectorizer.joblib')


def spam_or_ham():
    data_tup = import_subject()
    model = joblib.load('spam_detector.joblib')
    spam_senders_dict = {}
    for gmail, subject in data_tup:
        vector = cv.transform([subject])
        vector = vector.toarray()
        prediction = model.predict(vector)
        if (prediction == 1) and (gmail in spam_senders_dict):
            spam_senders_dict[gmail] += 1
        elif (prediction == 1) and (gmail not in spam_senders_dict):
            spam_senders_dict[gmail] = 1
    for gmail in spam_senders_dict:
        spam_senders.append(tuple([spam_senders_dict[gmail], gmail]))
    spam_senders.sort(reverse=True)


listener = sr.Recognizer()

engine = pyttsx3.init()
engine.setProperty('rate', 180)  # reduces WPM to 180
engine.setProperty('voice', 'com.apple.speech.synthesis.voice.samantha.premium')  # This voice is best suited in macos


def talk(text):
    engine.say(text)
    engine.runAndWait()


def mike_out():
    with sr.Microphone() as source:
        print("Adjusting noise ")
        listener.adjust_for_ambient_noise(source, duration=1)
        print("Listening..")
        recorded_audio = listener.listen(source)
        print("Done recording")
    try:
        info = listener.recognize_google(recorded_audio, language="en-US")
        print(info)
        return info.lower()
    except Exception as ex:
        print(ex)


def send_email(receiver, subject, message):
    server = smtplib.SMTP('smtp.gmail.com', 587)  # to connect to server
    server.starttls()  # tells that it is an secure connection
    # Make sure to give app access in your Google account
    server.login('gojo.testing123@gmail.com', 'hellogojo')
    email = EmailMessage()
    email['From'] = 'gojo.testing123@gmail.com'
    email['To'] = receiver
    email['Subject'] = subject
    email.set_content(message)
    server.send_message(email)
    server.close()


def gather_and_send():
    subject0 = ""
    body0 = ""
    i = 0
    j = 0
    for subs in sub:
        if i == 0:
            subject0 = subs.capitalize()
            i += 1
        else:
            subject0 += ('. ' + subs.capitalize())
    for bodies in body:
        if j == 0:
            body0 = bodies.capitalize()
            j += 1
        else:
            body0 += ('. ' + bodies.capitalize())
    unique_addresses = list(set(addresses))
    for receiver in unique_addresses:
        send_email(receiver, subject0, body0)
    print('Email sent successfully to', *unique_addresses)


def new_contact():
    for name, email in new_contacts:
        contact_list[name] = email
        new_cell0 = sheet.cell(x + 1, 1)
        new_cell1 = sheet.cell(x + 1, 2)
        new_cell2 = sheet.cell(x + 1, 3)
        new_cell0.value = x - 1
        new_cell1.value = name
        new_cell2.value = email
        wb.save('contacts.xlsx')
    wb.close()


def import_subject():
    username = "gojo.testing123@gmail.com"
    password = "hellogojo"
    data_tup = []
    # create an IMAP4 class with SSL
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    # authenticate
    imap.login(username, password)
    status, messages = imap.select("INBOX")

    # number of top emails to fetch
    n = 25
    # total number of emails
    messages = int(messages[0])
    for i in range(messages, messages - n, -1):
        # fetch the email message by ID
        res, msg = imap.fetch(str(i), "(RFC822)")
        for response in msg:
            if isinstance(response, tuple):
                # parse a bytes email into a message object
                msg = email.message_from_bytes(response[1])
                # decode the email subject
                subject, encoding = decode_header(msg["Subject"])[0]
                if isinstance(subject, bytes):
                    # if it's a bytes, decode to str
                    subject = subject.decode(encoding)
                # decode email sender
                From, encoding = decode_header(msg.get("From"))[0]
                if isinstance(From, bytes):
                    From = From.decode(encoding)
                print("Subject:", subject)
                print("From:", From)
                data_tup.append(tuple([From, subject]))
    return data_tup


screen_helper = """
ScreenManager:
    MenuScreen:
    SelectScreen0:
    SpamSenders:
    SelectScreen:
    NCScreen:
    SubjectScreen:
    BodyScreen:
    EndScreen:

<MenuScreen>:
    name: 'menu'
    Image:
        source: 'bot_anim.gif'
        anim_delay: 0.03
        mipmap: True
        allow_stretch: True
        pos_hint: {'center_x':0.5,'center_y':0.5}
    MDCard:
        size_hint: None, None
        size: "180dp", "70dp"
        pos_hint: {"center_x": 0.33, "center_y": 0.89}
        canvas.before:
            Color:
                rgba: app.theme_cls.primary_color
            RoundedRectangle:
                size: self.size
                pos: self.pos
                radius: [10]
        FloatLayout:
            size: self.size
            pos: self.pos
            pos_hint: {"center_x": 0.33, "center_y": 0.89}
    MDLabel:
        id: head
        text: 'HOWDY'
        bold: True
        halign: 'center'
        bold: True
        font_size: '30sp'
        pos_hint: {'center_x':0.3,'center_y':0.89}
        color: 0.9, 0.9, 0.9, 1
    MDFillRoundFlatButton:
        text: 'BEGIN'
        pos_hint: {'center_x':0.8,'center_y':0.14}
        on_press: root.manager.current = 'select0' 

<SelectScreen0>:
    name: 'select0'
    MDCard:
        size_hint: None, None
        size: "260dp", "180dp"
        text: 'Spam Detection'
        pos_hint: {"center_x": 0.5, "center_y": 0.33}
        ripple_behavior: True
        on_press: root.manager.current = 'spam'
        canvas.before:
            Color:
                rgba: app.theme_cls.primary_color
            RoundedRectangle:
                size: self.size
                pos: self.pos
                radius: [10]
        FloatLayout:
            size: self.size
            pos: self.pos
            pos_hint: {"center_x": 0.5, "center_y": 0.33}
            elevation: 6
    MDLabel:
        text: 'SPAM'
        bold: True
        font_size: '30sp'
        pos_hint: {"center_x": 0.65, "center_y": 0.335}
        color: 1, 1, 1, 1 
    MDLabel:
        text: 'DETECTION'
        bold: True
        font_size: '30sp'
        pos_hint: {"center_x": 0.65, "center_y": 0.285}
        color: 1, 1, 1, 1      
    MDCard:
        size_hint: None, None
        size: "260dp", "180dp"
        text: 'Write Email'
        pos_hint: {"center_x": 0.5, "center_y": 0.66}
        ripple_behavior: True
        on_press: root.manager.current = 'select'
        canvas.before:
            Color:
                rgba: app.theme_cls.primary_color
            RoundedRectangle:
                size: self.size
                pos: self.pos
                radius: [10]
        FloatLayout:
            size: self.size
            pos: self.pos
            pos_hint: {"center_x": 0.5, "center_y": 0.66}
            elevation: 6
    MDLabel:
        text: 'WRITE EMAIL'
        bold: True
        font_size: '30sp'
        pos_hint: {"center_x": 0.65, "center_y": 0.63}
        color: 1, 1, 1, 1
        
<SpamSenders>:
    name: 'spam'
    BoxLayout:
        ScrollView:
            MDList:
                id: scroll
    MDFillRoundFlatButton:
        text: 'Send Mail'
        pos_hint: {'center_x':0.59,'center_y':0.05}
        md_bg_color: app.theme_cls.primary_dark
        elevation: 12
        on_press: root.mail_spammers()
    MDFillRoundFlatButton:
        text: 'Menu'
        pos_hint: {'center_x':0.84,'center_y':0.05}
        on_press: root.manager.current = 'select0'

<SelectScreen>:
    name: 'select'
    BoxLayout:
        ScrollView:
            MDList:
                id: scroll
    MDFillRoundFlatButton:
        text: 'Add contact'
        pos_hint: {'center_x':0.185,'center_y':0.05}
        on_press: root.manager.current = 'new_contact'
    MDFillRoundFlatButton:
        text: 'Record VoiceOver'
        pos_hint: {'center_x':0.55,'center_y':0.05}
        on_press: root.receiver_addresses()
    MDFillRoundFlatButton:
        text: 'Next'
        pos_hint: {'center_x':0.895,'center_y':0.05}
        md_bg_color: app.theme_cls.primary_dark
        on_press: root.manager.current = 'subject'

<NCScreen>:
    name: 'new_contact'
    MDLabel:
        id: head
        text: 'NEW CONTACT'
        bold: True
        halign: 'center'
        bold: True
        font_size: '20sp'
        pos_hint: {'center_y':0.65}
        color: app.theme_cls.primary_dark
    MDLabel:
        id: head
        text: 'Multiple contacts can be Saved'
        halign: 'center'
        pos_hint: {'center_y':0.6}
        color: 0.8, 0.8, 0.8, 1
    MDTextField:
        id: name
        hint_text: "Enter name"
        helper_text: "Enter contact name"
        helper_text_mode: "on_focus"
        icon_right_color: app.theme_cls.primary_color
        pos_hint:{'center_x': 0.5, 'center_y': 0.5}
        color_mode: 'custom'
        line_color_normal: app.theme_cls.primary_color
        line_color_focus: app.theme_cls.primary_color
        size_hint_x:None
        width:500
    MDTextField:
        id: address
        hint_text: "Enter email address"
        helper_text: "Click on save"
        helper_text_mode: "on_focus"
        icon_right_color: app.theme_cls.primary_color
        pos_hint:{'center_x': 0.5, 'center_y': 0.4}
        color_mode: 'custom'
        line_color_normal: app.theme_cls.primary_color
        line_color_focus: app.theme_cls.primary_color
        size_hint_x:None
        width:500
    MDFillRoundFlatButton:
        text: 'Save'
        pos_hint: {'center_x': 0.5, 'center_y': 0.3}
        on_release: root.save_data()
    MDFillRoundFlatButton:
        text: 'Next'
        pos_hint: {'center_x': 0.88, 'center_y': 0.05}
        md_bg_color: app.theme_cls.primary_dark
        on_release: root.manager.current = 'subject'

<SubjectScreen>:
    name: 'subject'
    MDLabel:
        text: 'SUBJECT:'
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.148,'center_y':0.95}
        color: app.theme_cls.primary_dark
    MDRoundFlatButton:
        id: sub
        text: "Click on 'Record VoiceOver' generate Subject"
        pos_hint: {'center_x':0.5,'center_y':0.5}
        size_hint: 0.92, 0.82
    MDFillRoundFlatButton:
        text: 'Record VoiceOver'
        pos_hint: {'center_x':0.74,'center_y':0.95}
        md_bg_color: app.theme_cls.primary_dark
        on_press: root.listen_subject()
    MDFillRoundFlatButton:
        text: 'Next'
        pos_hint: {'center_x':0.87,'center_y':0.05}
        on_press: root.manager.current = 'body'

<BodyScreen>:
    name: 'body'
    MDLabel:
        text: 'EMAIL BODY:'
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.2,'center_y':0.95}
        color: app.theme_cls.primary_dark
    MDRoundFlatButton:
        id: ebod
        text: "Click on 'Record VoiceOver' generate Body"
        pos_hint: {'center_x':0.5,'center_y':0.5}
        size_hint: 0.92, 0.82
    MDFillRoundFlatButton:
        text: 'Record VoiceOver'
        pos_hint: {'center_x':0.74,'center_y':0.95}
        md_bg_color: app.theme_cls.primary_dark
        on_press: root.listen_body()
    MDFillRoundFlatButton:
        text: 'Next'
        pos_hint: {'center_x':0.87,'center_y':0.05}
        on_press: root.manager.current = 'end'

<EndScreen>:
    name: 'end'
    MDLabel:
        text: "Click on 'Send Mail'"
        pos_hint: {'center_x':0.64,'center_y':0.86}
        color: 0.8, 0.8, 0.8, 1
    MDFillRoundFlatButton:
        text: 'Send Mail'
        pos_hint: {'center_x':0.7,'center_y':0.86}
        on_press: root.final_send()
    MDCard:
        size_hint: None, None
        size: "260dp", "320dp"
        pos_hint: {"center_x": 0.5, "center_y": 0.5}
        canvas.before:
            Color:
                rgba: app.theme_cls.primary_color
            RoundedRectangle:
                size: self.size
                pos: self.pos
                radius: [10]
        FloatLayout:
            size: self.size
            pos: self.pos
            pos_hint: {"center_x": 0.5, "center_y": 0.5}
            elevation: 6
    MDLabel:
        text: 'Email(s) successfully sent to:'
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.5,'center_y':0.7}
        color: 1, 1, 1, 1
    MDLabel:
        id: final
        bold: True
        halign: 'center'
        pos_hint: {'center_x':0.5,'center_y':0.625}
        color: 1, 1, 1, 1
    MDFillRoundFlatButton:
        text: 'Menu'
        pos_hint: {'center_x':0.84,'center_y':0.05}
        on_press: root.manager.current = 'select0'
        
"""


class MenuScreen(Screen):
    pass


class SelectScreen0(Screen):
    pass


class SpamSenders(Screen):
    spam_or_ham()

    def mail_spammers(self):
        addresses.append("gojo.testing123@gmail.com")
        sub.append("Email spam senders!!")
        spam_mailbod = "Spam mail senders are:\n"
        i = 1
        for spam_num, name_address in spam_senders:
            if '<' in name_address:
                name_address = name_address.split("<")
                spam_mailbod += str(i) + ". Email address- " + (name_address[1])[:-1] + "\tName- " + name_address[0]
                spam_mailbod += "\nNo.of Spams sent- " + str(spam_num) + "\n"
                i += 1
            else:
                spam_mailbod += str(i) + ". Email address- " + name_address
                spam_mailbod += "\nNo.of Spams sent- " + str(spam_num) + "\n"
        body.append(spam_mailbod)
        gather_and_send()
        email_receivers.clear()
        addresses.clear()
        sub.clear()
        body.clear()


class SelectScreen(Screen):

    def receiver_addresses(self):
        while 1:
            receivers_str = mike_out()
            receivers = receivers_str.split(' and ')
            for receiver in receivers:
                if receiver not in contact_list:
                    receivers.remove(receiver)
            if len(receivers) == 0:
                talk("sorry, there are no similar email addresses in your contacts")
                talk("Please try again")
            else:
                break

        for receiver in receivers:
            email_receivers.append(receiver.capitalize())
            addresses.append(contact_list[receiver])
        print("Receiver's Email addresses: ", *addresses)


class NCScreen(Screen):

    def save_data(self):
        new_contacts.append([self.ids.name.text.lower(), self.ids.address.text.lower()])
        addresses.append(self.ids.address.text.lower().capitalize())
        email_receivers.append([self.ids.name.text.lower().capitalize()])
        new_contact()


class SubjectScreen(Screen):
    line_count = 0

    def listen_subject(self):
        sub_line = mike_out()
        sub.append(sub_line)
        dup = copy.deepcopy(sub_line)
        c = 0
        for index in range(len(dup) - 1):
            if dup[index] == " " and c < 7:
                c += 1
            elif dup[index] == " " and c >= 7:
                dup = dup[:index] + '\n' + dup[index + 1:]
                c = 0
        if self.line_count == 0:
            self.ids.sub.text = dup.capitalize()
            self.line_count += 1
        else:
            self.ids.sub.text += f"\n{dup.capitalize()}"


class BodyScreen(Screen):
    line_count = 0

    def listen_body(self):
        body_line = mike_out()
        body.append(body_line)
        dup = copy.deepcopy(body_line)
        c = 0
        for index in range(len(dup) - 1):
            if dup[index] == " " and c < 7:
                c += 1
            elif dup[index] == " " and c >= 7:
                dup = dup[:index] + '\n' + dup[index + 1:]
                c = 0
        if self.line_count == 0:
            self.ids.ebod.text = dup.capitalize()
            self.line_count += 1
        else:
            self.ids.ebod.text += f"\n{dup.capitalize()}"


class EndScreen(Screen):

    def final_send(self):
        gather_and_send()
        final_msg_names = ""
        unique_receivers = list(set(email_receivers))
        for receiver in unique_receivers:
            final_msg_names += (receiver + '\n')
        self.ids.final.text = final_msg_names
        email_receivers.clear()
        addresses.clear()
        sub.clear() 
        body.clear()


# Create the screen manager
sm = ScreenManager()
sm.add_widget(MenuScreen(name='menu'))
sm.add_widget(SelectScreen(name='select0'))
sm.add_widget(SelectScreen(name='spam'))
sm.add_widget(SelectScreen(name='select'))
sm.add_widget(NCScreen(name='new_contact'))
sm.add_widget(SubjectScreen(name='subject'))
sm.add_widget(BodyScreen(name='body'))
sm.add_widget(EndScreen(name='end'))


Window.size = (330, 600)


class DemoApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Blue"
        screen = Screen()
        self.help_str = Builder.load_string(screen_helper)
        screen.add_widget(self.help_str)
        for key in contact_list:
            icons = IconLeftWidget(
                icon=f"/Users/gurusaishreeshtirumalla/Desktop/Emailbot-ML/contact_icons/{contact_icons[key[0]]}")
            items = TwoLineAvatarListItem(text=key.capitalize(), secondary_text=contact_list[key])
            items.add_widget(icons)
            self.help_str.get_screen('select').ids.scroll.add_widget(items)
        for spam_num, address in spam_senders:
            icons = IconLeftWidget(
                icon=f"/Users/gurusaishreeshtirumalla/Desktop/Emailbot-ML/contact_icons/{contact_icons[address[0].lower()]}")
            items = OneLineAvatarIconListItem(text=(address.split("<"))[0])
            items.add_widget(icons)
            self.help_str.get_screen('spam').ids.scroll.add_widget(items)
        return screen


DemoApp().run()
